#!/usr/bin/env python3
# coding: utf8

import openpyxl
import datetime
import csv
import os
import re



def read_input_files_report(file_list):
    result = {}
    for file in file_list:
        print('Обработка файла: {}'.format(file))
        key_date = re.search(r'.*Край (.+)\.', file).group(1)
        wb = openpyxl.load_workbook(file)
        sh = wb.active
        result[key_date] = {}
        for row in range(6, sh.max_row + 1):
            if sh['H{}'.format(row)].value is None:
                continue
            address = re.search(r'((.+,){3})', sh['H{}'.format(row)].value.strip()).group(1)[:-1].replace(' ул.', '').replace(' пер.', '').replace(' сад.тов', '').replace(' мкр', '').replace(' скв', '').lower()
            result[key_date][address] = {'min_rate': sh['N{}'.format(row)].value, 'avg_rate': sh['O{}'.format(row)].value, 'session_count': sh['M{}'.format(row)].value, 'tariff': sh['K{}'.format(row)].value}
    return result


def read_input_argus(file_list):
    result = []
    for f in file_list:
        with open(f, encoding='windows-1251') as file:
            reader = csv.reader(file, delimiter=';')
            for row in reader:
                if (row[8] == 'DSL') and (row[24] == '1') and (row[84] == 'Ставропольский'):
                    open_date = datetime.datetime.strptime(re.search(r'(\d+\.\d+\.\d+) ', row[15]).group(1).replace('.', '-'), '%d-%m-%Y')
                    close_date = datetime.datetime.strptime(re.search(r'(\d+\.\d+\.\d+) ', row[16]).group(1).replace('.', '-'), '%d-%m-%Y')
                    address = '{}, {}, {}'.format(row[1], row[5], row[6])
                    address_format = address.replace(' корп.', 'к').lower()
                    result.append({'number': row[13], 'open_date': open_date, 'close_date': close_date,'address': address, 'address_format': address_format})                    
    return result    


def main():
    while True:
        try:            
            start_date = datetime.datetime.strptime(input('Начальная дата (д/м/г): '), '%d/%m/%y')
            end_date = datetime.datetime.strptime(input('Конечная дата (д/м/г): '), '%d/%m/%y') #+ datetime.timedelta(minutes=1439)
        except:
            continue
        else:
            break

    file_list = ['input{}report'.format(os.sep) + os.sep + x for x in os.listdir('input{}report'.format(os.sep))]
    data_report = read_input_files_report(file_list)
        
    file_list = ['input{}argus'.format(os.sep) + os.sep + x for x in os.listdir('input{}argus'.format(os.sep))]
    data_argus = read_input_argus(file_list)
    
    wb = openpyxl.load_workbook('resource{}template.xlsx'.format(os.sep))
    sh = wb.active
    row_out = 9
    bad = 0
    ok = 0
    count = 0
    
    
    for _ in data_argus:
        bad_day = 0
        number = _['number']
        open_date = _['open_date']
        close_date = _['close_date']
        address = _['address']
        address_format = _['address_format']

        if (close_date < start_date) or (close_date > end_date):
            continue
        
        count += 1

        d1 = (open_date - datetime.timedelta(1)).strftime('%Y-%m-%d')
        d2 = (open_date - datetime.timedelta(2)).strftime('%Y-%m-%d')
        d3 = (open_date - datetime.timedelta(3)).strftime('%Y-%m-%d')

        # -1 день
        try:
            sh['F{}'.format(row_out)].value = data_report[d1][address_format]['min_rate']
            sh['G{}'.format(row_out)].value = data_report[d1][address_format]['avg_rate']
            sh['H{}'.format(row_out)].value = data_report[d1][address_format]['session_count']
            sh['E{}'.format(row_out)].value = data_report[d1][address_format]['tariff']
        except:
            bad_day += 1
        # -2 дня
        try:
            sh['I{}'.format(row_out)].value = data_report[d2][address_format]['min_rate']
            sh['J{}'.format(row_out)].value = data_report[d2][address_format]['avg_rate']
            sh['K{}'.format(row_out)].value = data_report[d2][address_format]['session_count']
            sh['E{}'.format(row_out)].value = data_report[d2][address_format]['tariff']
        except:
            bad_day += 1
        # -3 дня
        try:
            sh['L{}'.format(row_out)].value = data_report[d3][address_format]['min_rate']
            sh['M{}'.format(row_out)].value = data_report[d3][address_format]['avg_rate']
            sh['N{}'.format(row_out)].value = data_report[d3][address_format]['session_count']
            sh['E{}'.format(row_out)].value = data_report[d3][address_format]['tariff']
        except:
            bad_day += 1
            
        if bad_day == 3:
            bad += 1
            print(number, open_date.strftime('%d-%m-%y'), address)
            continue
        sh['B{}'.format(row_out)].value = number
        sh['C{}'.format(row_out)].value = address
        sh['D{}'.format(row_out)].value = open_date.strftime('%d-%m-%y')
        row_out += 1
        ok += 1
    sh['C2'].value = count
    sh['C3'].value = ok
    sh['C4'].value = bad
    wb.save('out{}{}-{}.xlsx'.format(os.sep, start_date.strftime('%d.%m.%y'), end_date.strftime('%d.%m.%y')))
        

if __name__ == '__main__':
    cur_dir = os.sep.join(os.path.abspath(__file__).split(os.sep)[:-1])
    os.chdir(cur_dir)
    main()